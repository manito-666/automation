# coding:utf-8
import os,sys
curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = os.path.split(curPath)[0]
sys.path.append(rootPath)
from openpyxl import load_workbook

data_debug='/Users/wangzhipeng/PycharmProjects/automation/util/data/result.xlsx'

'''读取excel表格数据'''
def read_data(sheet_name,path,case_id):
    workbook1 = load_workbook(path)
    # 定位表单（test_data）
    sheet1 = workbook1[sheet_name]
    test_case = []  # 用来存储每一行数据，也就是一条测试用例
    test_case.append(sheet1.cell(case_id+1,1).value)
    test_case.append(sheet1.cell(case_id+1,2).value)
    test_case.append(sheet1.cell(case_id+1,3).value)
    test_case.append(sheet1.cell(case_id+1,4).value)
    test_case.append(sheet1.cell(case_id+1,5).value)
    test_case.append(sheet1.cell(case_id+1,6).value)
    test_case.append(sheet1.cell(case_id + 1, 8).value)
    test_case.append(sheet1.cell(case_id + 1, 9).value)
    test_case.append(sheet1.cell(case_id + 1, 10).value)
    test_case.append(sheet1.cell(case_id + 1, 7).value)
    test_case.append(sheet1.cell(case_id + 1, 11).value)
    test_case.append(sheet1.cell(case_id + 1, 12).value)
    # test_case.append(sheet1.cell(case_id + 1, 13).value)
    return test_case


# 将测试结果写excel
def write_data(sheet_name, row, col, value):
    workbook1 = load_workbook(data_debug)
    sheet = workbook1[sheet_name]
    sheet.cell(row, col).value = value
    workbook1.save(data_debug)

# 统计测试用例的行数
def count_case(sheet_name,path):
    workbook1 = load_workbook(path)
    sheet = workbook1[sheet_name]
    max_row = sheet.max_row  # 统计测试用例的行数
    return max_row



if __name__ == '__main__':
    read_data('sheet', 'result.xlsx', 1)
